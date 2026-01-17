
"""
Indian Stock Market Analysis Dashboard
Analysis of Nifty 50 Growth: Revenue Expansion vs Margin Re-Rating
Version 2.4.0 - Restored Original State (7 Pages)
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import numpy as np
from datetime import datetime

from config import COLORS, AUTHOR, BRAND_NAME, EXPERIENCE, LOCATION, YEAR, PAGES
from data import generate_data
from styles import (
    get_custom_css, display_styled_dataframe,
    render_section_header, render_subsection_header, render_divider,
    render_info_box, render_warning_box, render_success_box,
    render_footer
)

# ═══════════════════════════════════════════════════════════════════════════
# PAGE CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Indian Stock Market Analysis",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown(get_custom_css(), unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════
# SIDEBAR & NAVIGATION
# ═══════════════════════════════════════════════════════════════════════════

st.sidebar.markdown(f"# 📊 {BRAND_NAME}")
st.sidebar.markdown("---")
st.sidebar.markdown(f"**Prof. V. Ravichandran**")
st.sidebar.markdown(f"*{EXPERIENCE}*")
st.sidebar.markdown("---")

# Only show first 7 pages (exclude data sources)
pages_list = PAGES[:7]

page = st.sidebar.radio(
    "📍 Choose Analysis:",
    pages_list,
    key="main_nav"
)

st.sidebar.markdown("---")
st.sidebar.markdown(f"📍 {LOCATION} | {YEAR}")

# ═══════════════════════════════════════════════════════════════════════════
# DOWNLOAD EXCEL FEATURE
# ═══════════════════════════════════════════════════════════════════════════

st.sidebar.markdown("---")
st.sidebar.markdown("### 📥 Download Report")

if st.sidebar.button("📊 Download Analysis (Excel)", use_container_width=True):
    # Import openpyxl for Excel creation
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Get all data
    five_year = data['five_year']
    quarterly = data['quarterly']
    sectors = data['sector']
    downgrades = data['downgrades']
    scenarios = data['scenarios']
    nifty_levels = data['nifty_levels']
    metrics = data['metrics']
    
    # Define styles
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    title_font = Font(bold=True, color="000000", size=14)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Executive Summary
    ws_summary = wb.create_sheet("Executive Summary")
    ws_summary['A1'] = "Indian Stock Market Analysis Dashboard"
    ws_summary['A1'].font = title_font
    ws_summary['A1'].fill = title_fill
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A3'] = "Key Metrics"
    ws_summary['A3'].font = header_font
    ws_summary['A3'].fill = header_fill
    
    row = 4
    for key, value in metrics.items():
        ws_summary[f'A{row}'] = key.replace('_', ' ').title()
        ws_summary[f'B{row}'] = value
        row += 1
    
    # Sheet 2: 5-Year Analysis
    ws_5year = wb.create_sheet("5-Year Trend")
    ws_5year['A1'] = "5-Year Performance Analysis"
    ws_5year['A1'].font = title_font
    ws_5year['A1'].fill = title_fill
    ws_5year.merge_cells('A1:F1')
    
    # Headers
    for col, header in enumerate(five_year.columns, 1):
        cell = ws_5year.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Data
    for r_idx, row_data in enumerate(five_year.values, 4):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws_5year.cell(row=r_idx, column=c_idx)
            cell.value = value
            cell.border = thin_border
    
    # Sheet 3: Quarterly Analysis
    ws_quarterly = wb.create_sheet("Quarterly Analysis")
    ws_quarterly['A1'] = "FY2025 Quarterly Performance"
    ws_quarterly['A1'].font = title_font
    ws_quarterly['A1'].fill = title_fill
    ws_quarterly.merge_cells('A1:D1')
    
    # Headers
    for col, header in enumerate(quarterly.columns, 1):
        cell = ws_quarterly.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Data
    for r_idx, row_data in enumerate(quarterly.values, 4):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws_quarterly.cell(row=r_idx, column=c_idx)
            cell.value = value
            cell.border = thin_border
    
    # Sheet 4: Sector Analysis
    ws_sector = wb.create_sheet("Sector Analysis")
    ws_sector['A1'] = "Sector Performance"
    ws_sector['A1'].font = title_font
    ws_sector['A1'].fill = title_fill
    ws_sector.merge_cells('A1:E1')
    
    # Headers
    for col, header in enumerate(sectors.columns, 1):
        cell = ws_sector.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Data
    for r_idx, row_data in enumerate(sectors.values, 4):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws_sector.cell(row=r_idx, column=c_idx)
            cell.value = value
            cell.border = thin_border
    
    # Sheet 5: Earnings Downgrades
    ws_downgrades = wb.create_sheet("Earnings Downgrades")
    ws_downgrades['A1'] = "6-Month Earnings Revision Trend"
    ws_downgrades['A1'].font = title_font
    ws_downgrades['A1'].fill = title_fill
    ws_downgrades.merge_cells('A1:C1')
    
    # Headers
    for col, header in enumerate(downgrades.columns, 1):
        cell = ws_downgrades.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Data
    for r_idx, row_data in enumerate(downgrades.values, 4):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws_downgrades.cell(row=r_idx, column=c_idx)
            cell.value = value
            cell.border = thin_border
    
    # Sheet 6: Scenarios
    ws_scenarios = wb.create_sheet("Investment Scenarios")
    ws_scenarios['A1'] = "Investment Scenarios Analysis"
    ws_scenarios['A1'].font = title_font
    ws_scenarios['A1'].fill = title_fill
    ws_scenarios.merge_cells('A1:D1')
    
    row = 3
    for scenario_name, scenario_data in scenarios.items():
        ws_scenarios[f'A{row}'] = scenario_name
        ws_scenarios[f'A{row}'].font = Font(bold=True, size=11)
        row += 1
        
        ws_scenarios[f'A{row}'] = "Description"
        ws_scenarios[f'B{row}'] = scenario_data['description']
        row += 1
        
        ws_scenarios[f'A{row}'] = "Probability"
        ws_scenarios[f'B{row}'] = f"{scenario_data['probability']*100:.0f}%"
        row += 1
        
        ws_scenarios[f'A{row}'] = "FY25 Earnings"
        ws_scenarios[f'B{row}'] = scenario_data['fy25_earnings']
        row += 1
        
        ws_scenarios[f'A{row}'] = "FY26 Earnings"
        ws_scenarios[f'B{row}'] = scenario_data['fy26_earnings']
        row += 1
        
        ws_scenarios[f'A{row}'] = "FY27 Earnings"
        ws_scenarios[f'B{row}'] = scenario_data['fy27_earnings']
        row += 1
        
        ws_scenarios[f'A{row}'] = "FY25 P/E"
        ws_scenarios[f'B{row}'] = scenario_data['fy25_pe']
        row += 1
        
        ws_scenarios[f'A{row}'] = "FY26 P/E"
        ws_scenarios[f'B{row}'] = scenario_data['fy26_pe']
        row += 1
        
        ws_scenarios[f'A{row}'] = "FY27 P/E"
        ws_scenarios[f'B{row}'] = scenario_data['fy27_pe']
        row += 1
        
        ws_scenarios[f'A{row}'] = "Nifty Targets"
        ws_scenarios[f'B{row}'] = f"FY25: {nifty_levels[scenario_name][0]:.0f} | FY26: {nifty_levels[scenario_name][1]:.0f} | FY27: {nifty_levels[scenario_name][2]:.0f}"
        row += 2
    
    # Auto-adjust column widths
    for ws in [ws_summary, ws_5year, ws_quarterly, ws_sector, ws_downgrades, ws_scenarios]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Download button
    st.sidebar.download_button(
        label="⬇️ Click to Download",
        data=excel_buffer,
        file_name=f"Nifty_50_Analysis_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

st.sidebar.markdown("---")

# ═══════════════════════════════════════════════════════════════════════════
# LOAD DATA
# ═══════════════════════════════════════════════════════════════════════════

@st.cache_data
def load_dashboard_data():
    return generate_data()

data = load_dashboard_data()

# ═══════════════════════════════════════════════════════════════════════════
# PAGE 0: OVERVIEW
# ═══════════════════════════════════════════════════════════════════════════

if page == PAGES[0]:
    render_section_header("📈 Nifty 50 Analysis: Growth Drivers & Sustainability")
    
    st.markdown("""
    **Analysis Period:** FY2021 - FY2025 YTD  
    **Focus:** Is growth driven by revenue expansion or margin re-rating?
    """)
    
    render_divider()
    
    # Key Metrics
    render_subsection_header("📊 Key Metrics Summary")
    
    metrics_data = {
        'Metric': ['Revenue CAGR (FY21-25)', 'Profit CAGR (FY21-25)', 'EBITDA Margin (FY25)', 'PAT Margin (FY25)'],
        'Value': ['9.2%', '19.8%', '33.0%', '10.5%']
    }
    
    display_styled_dataframe(
        pd.DataFrame(metrics_data),
        columns_to_style=['Value'],
        width='stretch'
    )
    
    render_divider()
    
    # Key Insights
    render_subsection_header("📌 Key Insights")
    
    col1, col2 = st.columns(2)
    
    with col1:
        render_success_box(
            "**FY2021-2024: Healthy Growth**\n\n"
            "✅ Revenue expanding (+10.5% avg)\n"
            "✅ Margins improving (+50 bps)\n"
            "✅ Both drivers working\n"
            "✅ Sustainable model"
        )
    
    with col2:
        render_warning_box(
            "**FY2025: Concerning Shift**\n\n"
            "⚠️ Revenue decelerating (6.9%)\n"
            "⚠️ Margins propping profits\n"
            "⚠️ One-time tailwinds fading\n"
            "❌ Unsustainable"
        )
    
    render_divider()
    
    # Investment Takeaway
    render_info_box(
        "**Investment Perspective**\n\n"
        "The Nifty 50 profit growth (19.8% CAGR) masks slowing revenue growth (9.2% CAGR). "
        "While margin expansion provided tailwinds through FY24, FY25 shows concerning revenue deceleration. "
        "Investors should focus on revenue growth sustainability rather than margin-driven earnings growth."
    )

# ═══════════════════════════════════════════════════════════════════════════
# PAGE 1: 5-YEAR TREND
# ═══════════════════════════════════════════════════════════════════════════

elif page == PAGES[1]:
    render_section_header("📈 5-Year Trend Analysis")
    
    render_subsection_header("💹 5-Year Performance")
    
    five_year = data['five_year']
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=five_year['Fiscal Year'],
        y=five_year['Revenue Growth (%)'],
        mode='lines+markers',
        name='Revenue Growth',
        line=dict(color=COLORS['chart_blue'], width=3),
        marker=dict(size=10)
    ))
    
    fig.add_trace(go.Scatter(
        x=five_year['Fiscal Year'],
        y=five_year['PAT Growth (%)'],
        mode='lines+markers',
        name='Profit Growth',
        line=dict(color=COLORS['accent_red'], width=3),
        marker=dict(size=10)
    ))
    
    fig.update_layout(
        title="Revenue vs Profit Growth Trends",
        xaxis_title="Fiscal Year",
        yaxis_title="Growth Rate (%)",
        template='plotly_white',
        height=400,
        hovermode='x unified'
    )
    
    st.plotly_chart(fig, width='stretch')
    
    render_divider()
    
    render_subsection_header("📊 Margin Trends")
    
    fig2 = go.Figure()
    
    fig2.add_trace(go.Scatter(
        x=five_year['Fiscal Year'],
        y=five_year['EBITDA Margin (%)'],
        mode='lines+markers',
        name='EBITDA Margin',
        line=dict(color=COLORS['accent_gold'], width=3),
        marker=dict(size=10)
    ))
    
    fig2.add_trace(go.Scatter(
        x=five_year['Fiscal Year'],
        y=five_year['PAT Margin (%)'],
        mode='lines+markers',
        name='PAT Margin',
        line=dict(color=COLORS['accent_red'], width=3),
        marker=dict(size=10)
    ))
    
    fig2.update_layout(
        title="Margin Trends",
        xaxis_title="Fiscal Year",
        yaxis_title="Margin (%)",
        template='plotly_white',
        height=400,
        hovermode='x unified'
    )
    
    st.plotly_chart(fig2, width='stretch')
    
    render_divider()
    
    display_styled_dataframe(
        five_year,
        columns_to_style=['Revenue Growth (%)', 'EBITDA Growth (%)', 'PAT Growth (%)'],
        width='stretch',
        hide_index=True
    )

# ═══════════════════════════════════════════════════════════════════════════
# PAGE 2: QUARTERLY DEEP-DIVE
# ═══════════════════════════════════════════════════════════════════════════

elif page == PAGES[2]:
    render_section_header("📊 FY2025 Quarterly Deep-Dive Analysis")
    
    st.markdown("""
    **Comprehensive Annual vs Quarterly Performance Analysis**  
    Analyzing FY2025 performance through both annual and quarterly lens
    """)
    
    render_divider()
    
    # Get data
    five_year = data['five_year']
    quarterly = data['quarterly']
    
    # ANNUAL PERFORMANCE (Last row of 5-year data)
    render_subsection_header("📈 Annual Performance (FY2025 YTD)")
    
    annual_row = five_year.iloc[-1]
    annual_cols = st.columns(4)
    with annual_cols[0]:
        st.metric("Revenue Growth", f"{annual_row['Revenue Growth (%)']:.1f}%", delta="YoY")
    with annual_cols[1]:
        st.metric("EBITDA Growth", f"{annual_row['EBITDA Growth (%)']:.1f}%", delta="YoY")
    with annual_cols[2]:
        st.metric("PAT Growth", f"{annual_row['PAT Growth (%)']:.1f}%", delta="YoY")
    with annual_cols[3]:
        st.metric("EBITDA Margin", f"{annual_row['EBITDA Margin (%)']:.1f}%", delta="vs FY24")
    
    render_divider()
    
    # QUARTERLY PERFORMANCE
    render_subsection_header("📊 Quarterly Breakdown (FY2025)")
    
    display_styled_dataframe(
        quarterly,
        columns_to_style=['Revenue Growth (%)', 'EBITDA Growth (%)', 'PAT Growth (%)'],
        width='stretch',
        hide_index=True
    )
    
    render_divider()
    
    # QUARTERLY vs ANNUAL COMPARISON
    render_subsection_header("🔍 Quarterly vs Annual Comparison")
    
    col1, col2 = st.columns(2)
    
    with col1:
        render_info_box(
            "**Quarterly Trend**\n\n"
            "• Q1 FY25: Revenue 9.6% | PAT 0.8%\n"
            "• Q2 FY25: Revenue 6.6% | PAT -1.0%\n"
            "• Q3 FY25: Revenue 4.5% | PAT 9.5%\n\n"
            "**Observation:** Revenue declining QoQ while profit recovery in Q3"
        )
    
    with col2:
        render_warning_box(
            "**Key Insights**\n\n"
            "• Annual revenue growth (6.9%) lower than prior years\n"
            "• Q2 showed negative PAT growth (-1.0%)\n"
            "• Q3 recovery driven by margin expansion\n"
            "• Divergence between revenue and profit trends"
        )
    
    render_divider()
    
    # COMPARATIVE CHART
    render_subsection_header("📈 Annual Trend with Quarterly Details")
    
    fig = go.Figure()
    
    # Annual trend (5-year)
    fig.add_trace(go.Scatter(
        x=five_year['Fiscal Year'],
        y=five_year['Revenue Growth (%)'],
        mode='lines+markers',
        name='Annual Revenue Growth',
        line=dict(color=COLORS['chart_blue'], width=3, dash='solid'),
        marker=dict(size=10)
    ))
    
    # Quarterly points
    q_labels = quarterly['Quarter'].tolist()
    fig.add_trace(go.Scatter(
        x=q_labels,
        y=quarterly['Revenue Growth (%)'],
        mode='markers',
        name='Quarterly Revenue Growth',
        marker=dict(size=12, color=COLORS['accent_red'], symbol='diamond')
    ))
    
    fig.update_layout(
        title="Annual vs Quarterly Revenue Growth Trend",
        xaxis_title="Period",
        yaxis_title="Growth Rate (%)",
        template='plotly_white',
        height=400,
        hovermode='x unified'
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    render_divider()
    
    render_info_box(
        "**Deep-Dive Takeaway**\n\n"
        "FY2025 shows concerning revenue deceleration when analyzed quarterly, suggesting temporary tailwinds may have faded. "
        "However, profit recovery in Q3 indicates management's ability to defend margins through cost management. "
        "The key question: Is Q3 recovery sustainable, or is it driven by one-time factors?"
    )

# ═══════════════════════════════════════════════════════════════════════════
# PAGE 3: SECTOR ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════

elif page == PAGES[3]:
    render_section_header("🏦 Sector Performance Analysis")
    
    sectors = data['sector']
    display_styled_dataframe(
        sectors,
        columns_to_style=['Contribution (%)', 'Growth (%)'],
        width='stretch',
        hide_index=True
    )

# ═══════════════════════════════════════════════════════════════════════════
# PAGE 4: EARNINGS DOWNGRADES
# ═══════════════════════════════════════════════════════════════════════════

elif page == PAGES[4]:
    render_section_header("📉 6-Month Earnings Revision Trend")
    
    downgrades = data['downgrades']
    display_styled_dataframe(
        downgrades,
        columns_to_style=['Direction (%)'],
        width='stretch',
        hide_index=True
    )

# ═══════════════════════════════════════════════════════════════════════════
# PAGE 5: SCENARIOS
# ═══════════════════════════════════════════════════════════════════════════

elif page == PAGES[5]:
    render_section_header("🎯 Investment Scenarios - Detailed Analysis")
    
    st.markdown("""
    **Select a scenario below to view detailed analysis including:**
    - Earnings projections (FY2025-2027)
    - P/E multiple assumptions
    - Nifty 50 target levels
    - Probability-weighted returns
    """)
    
    render_divider()
    
    # Get scenarios data
    scenarios_data = data['scenarios']
    nifty_levels = data['nifty_levels']
    
    # Radio button to select scenario
    scenario_names = list(scenarios_data.keys())
    selected_scenario = st.radio(
        "📍 Choose Investment Scenario:",
        scenario_names,
        index=0,
        key="scenario_selector"
    )
    
    render_divider()
    
    # Get selected scenario data
    scenario_info = scenarios_data[selected_scenario]
    nifty_targets = nifty_levels[selected_scenario]
    
    # Display selected scenario
    render_subsection_header(f"📊 {selected_scenario}")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown(f"""
        **Scenario Description:**
        
        {scenario_info['description']}
        
        **Probability:** {scenario_info['probability']*100:.0f}%
        """)
    
    with col2:
        # Color indicator with HTML rendering
        st.markdown(f"""
        **Scenario Type & Color:**
        """)
        
        # Display colored indicator
        color = scenario_info['color']
        st.markdown(f"<p style='font-size: 24px; color: {color};'>● {selected_scenario}</p>", unsafe_allow_html=True)
        
        st.markdown(f"""
        **Key Characteristics:**
        
        {scenario_info['description']}
        """)
    
    render_divider()
    
    # Earnings Projections
    render_subsection_header("💰 Earnings Projections (FY2025-2027)")
    
    earnings_col1, earnings_col2, earnings_col3 = st.columns(3)
    
    with earnings_col1:
        st.metric("FY2025 Earnings", f"₹{scenario_info['fy25_earnings']:.1f}", delta="Growth")
    with earnings_col2:
        st.metric("FY2026 Earnings", f"₹{scenario_info['fy26_earnings']:.1f}", delta="CAGR")
    with earnings_col3:
        st.metric("FY2027 Earnings", f"₹{scenario_info['fy27_earnings']:.1f}", delta="Projection")
    
    render_divider()
    
    # P/E Multiples
    render_subsection_header("📈 P/E Multiple Assumptions")
    
    pe_col1, pe_col2, pe_col3 = st.columns(3)
    
    with pe_col1:
        st.metric("FY2025 P/E", f"{scenario_info['fy25_pe']:.1f}x", delta="Valuation")
    with pe_col2:
        st.metric("FY2026 P/E", f"{scenario_info['fy26_pe']:.1f}x", delta="Normalized")
    with pe_col3:
        st.metric("FY2027 P/E", f"{scenario_info['fy27_pe']:.1f}x", delta="Terminal")
    
    render_divider()
    
    # Nifty 50 Target Levels
    render_subsection_header("🎯 Nifty 50 Target Levels")
    
    target_col1, target_col2, target_col3 = st.columns(3)
    
    with target_col1:
        st.metric("FY2025 Target", f"{nifty_targets[0]:.0f}", delta="Near-term")
    with target_col2:
        st.metric("FY2026 Target", f"{nifty_targets[1]:.0f}", delta="Medium-term")
    with target_col3:
        st.metric("FY2027 Target", f"{nifty_targets[2]:.0f}", delta="Long-term")
    
    render_divider()
    
    # Scenario Analysis Table
    render_subsection_header("📊 Scenario Comparison Matrix")
    
    # Create comparison dataframe
    comparison_df = pd.DataFrame({
        'Metric': ['Probability', 'FY25 Earnings', 'FY26 Earnings', 'FY27 Earnings', 
                   'FY25 P/E', 'FY26 P/E', 'FY27 P/E',
                   'FY25 Target', 'FY26 Target', 'FY27 Target'],
        'Base Case (50%)': [
            f"{scenarios_data['Base Case (50%)']['probability']*100:.0f}%",
            f"₹{scenarios_data['Base Case (50%)']['fy25_earnings']:.1f}",
            f"₹{scenarios_data['Base Case (50%)']['fy26_earnings']:.1f}",
            f"₹{scenarios_data['Base Case (50%)']['fy27_earnings']:.1f}",
            f"{scenarios_data['Base Case (50%)']['fy25_pe']:.1f}x",
            f"{scenarios_data['Base Case (50%)']['fy26_pe']:.1f}x",
            f"{scenarios_data['Base Case (50%)']['fy27_pe']:.1f}x",
            f"{nifty_levels['Base Case (50%)'][0]:.0f}",
            f"{nifty_levels['Base Case (50%)'][1]:.0f}",
            f"{nifty_levels['Base Case (50%)'][2]:.0f}"
        ],
        'Bear Case (25%)': [
            f"{scenarios_data['Bear Case (25%)']['probability']*100:.0f}%",
            f"₹{scenarios_data['Bear Case (25%)']['fy25_earnings']:.1f}",
            f"₹{scenarios_data['Bear Case (25%)']['fy26_earnings']:.1f}",
            f"₹{scenarios_data['Bear Case (25%)']['fy27_earnings']:.1f}",
            f"{scenarios_data['Bear Case (25%)']['fy25_pe']:.1f}x",
            f"{scenarios_data['Bear Case (25%)']['fy26_pe']:.1f}x",
            f"{scenarios_data['Bear Case (25%)']['fy27_pe']:.1f}x",
            f"{nifty_levels['Bear Case (25%)'][0]:.0f}",
            f"{nifty_levels['Bear Case (25%)'][1]:.0f}",
            f"{nifty_levels['Bear Case (25%)'][2]:.0f}"
        ],
        'Bull Case (25%)': [
            f"{scenarios_data['Bull Case (25%)']['probability']*100:.0f}%",
            f"₹{scenarios_data['Bull Case (25%)']['fy25_earnings']:.1f}",
            f"₹{scenarios_data['Bull Case (25%)']['fy26_earnings']:.1f}",
            f"₹{scenarios_data['Bull Case (25%)']['fy27_earnings']:.1f}",
            f"{scenarios_data['Bull Case (25%)']['fy25_pe']:.1f}x",
            f"{scenarios_data['Bull Case (25%)']['fy26_pe']:.1f}x",
            f"{scenarios_data['Bull Case (25%)']['fy27_pe']:.1f}x",
            f"{nifty_levels['Bull Case (25%)'][0]:.0f}",
            f"{nifty_levels['Bull Case (25%)'][1]:.0f}",
            f"{nifty_levels['Bull Case (25%)'][2]:.0f}"
        ]
    })
    
    display_styled_dataframe(
        comparison_df,
        width='stretch',
        hide_index=True
    )
    
    render_divider()
    
    # Investment Perspective
    if selected_scenario == 'Base Case (50%)':
        render_success_box(
            "**Base Case (Most Likely - 50% Probability)**\n\n"
            "Margin resilience with slow revenue growth. Earnings grow from ₹5.5 (FY25) to ₹12.5 (FY27). "
            "P/E multiple compresses from 25x to 24x, limiting re-rating. Nifty target ranges from 56,700 to 67,900. "
            "This is the consensus scenario with moderate upside."
        )
    elif selected_scenario == 'Bear Case (25%)':
        render_warning_box(
            "**Bear Case (Stress - 25% Probability)**\n\n"
            "Margin compression due to input cost spike. Earnings growth severely impacted: ₹2.0 → ₹7.5. "
            "P/E multiple contracts from 23x to 21.5x. Nifty downside risk to 50,400-53,200. "
            "Triggered by commodities rally or demand shock."
        )
    else:
        render_info_box(
            "**Bull Case (Optimistic - 25% Probability)**\n\n"
            "Revenue recovery accelerates with margin stability. Strong earnings growth: ₹9.0 → ₹15.5. "
            "P/E multiple expands from 25.5x to 26.5x as confidence returns. Nifty upside to 59,700-81,700. "
            "Requires revenue inflection + operational efficiency."
        )

# ═══════════════════════════════════════════════════════════════════════════
# PAGE 6: DATA EXPLORER
# ═══════════════════════════════════════════════════════════════════════════

elif page == PAGES[6]:
    render_section_header("📋 Data Explorer")
    
    st.markdown("**All Datasets**")
    
    tab1, tab2, tab3, tab4 = st.tabs(["5-Year", "Quarterly", "Sectors", "Downgrades"])
    
    with tab1:
        display_styled_dataframe(data['five_year'], width='stretch', hide_index=True)
    
    with tab2:
        display_styled_dataframe(data['quarterly'], width='stretch', hide_index=True)
    
    with tab3:
        display_styled_dataframe(data['sector'], width='stretch', hide_index=True)
    
    with tab4:
        display_styled_dataframe(data['downgrades'], width='stretch', hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════
# FOOTER
# ═══════════════════════════════════════════════════════════════════════════

st.markdown("---")
render_footer(AUTHOR, BRAND_NAME, "NSE, RBI, BSE, MCA, SEBI | Research: Business Standard, Economic Times, Brokerages")
